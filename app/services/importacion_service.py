"""
Servicio para importación masiva de empresas desde Excel
"""
from typing import List, Dict, Any, Tuple
from datetime import datetime
import openpyxl
from openpyxl.workbook import Workbook
from io import BytesIO
from app.models.empresa import Empresa
from app.models.enums import EstadoEmpresa
from app.repositories.empresa_repository import EmpresaRepository


class ImportacionService:
    """
    Servicio para importar empresas desde archivos Excel
    """
    
    # Columnas esperadas en el Excel (orden)
    COLUMNAS_ESPERADAS = [
        'NIT',
        'RAZON_SOCIAL',
        'ESTADO',
        'CERTIFICADO_VENCIMIENTO',
        'CERTIFICADO_RENOVADO',
        'CERTIFICADO_FACTURADO',
        'RESOLUCION_VENCIMIENTO',
        'RESOLUCION_RENOVADO',
        'RESOLUCION_FACTURADO',
        'DOCUMENTO_VENCIMIENTO',
        'DOCUMENTO_RENOVADO',
        'DOCUMENTO_FACTURADO'
    ]
    
    def __init__(self, repository: EmpresaRepository):
        """
        Inicializa el servicio
        
        Args:
            repository: Repositorio de empresas
        """
        self.repository = repository
    
    def validar_estructura_excel(self, file_content: bytes) -> Dict[str, Any]:
        """
        Valida que el Excel tenga la estructura correcta
        
        Args:
            file_content: Contenido del archivo Excel
            
        Returns:
            Diccionario con el resultado de la validación
        """
        try:
            wb = openpyxl.load_workbook(BytesIO(file_content))
            ws = wb.active
            
            # Verificar que tenga encabezados
            if ws.max_row < 2:
                return {
                    'success': False,
                    'error': 'El archivo está vacío o no tiene datos'
                }
            
            # Obtener encabezados (primera fila)
            encabezados = [cell.value for cell in ws[1]]
            
            # Validar que tenga todas las columnas esperadas
            columnas_faltantes = []
            for columna in self.COLUMNAS_ESPERADAS:
                if columna not in encabezados:
                    columnas_faltantes.append(columna)
            
            if columnas_faltantes:
                return {
                    'success': False,
                    'error': f'Faltan las siguientes columnas: {", ".join(columnas_faltantes)}'
                }
            
            return {
                'success': True,
                'message': 'Estructura válida',
                'total_filas': ws.max_row - 1  # Restar encabezados
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Error al leer el archivo: {str(e)}'
            }
    
    def procesar_fila(self, fila: List[Any], numero_fila: int) -> Tuple[Empresa, List[str]]:
        """
        Procesa una fila del Excel y crea un objeto Empresa
        
        Args:
            fila: Lista con los valores de la fila
            numero_fila: Número de fila (para reportar errores)
            
        Returns:
            Tupla (Empresa o None, Lista de errores)
        """
        errores = []
        
        try:
            # Extraer datos
            nit = str(fila[0]).strip() if fila[0] else None
            nombre = str(fila[1]).strip() if fila[1] else None
            estado = str(fila[2]).strip().lower() if fila[2] else 'activo'
            
            # Validaciones básicas
            if not nit:
                errores.append(f'Fila {numero_fila}: NIT es obligatorio')
                return None, errores
            
            if not nombre:
                errores.append(f'Fila {numero_fila}: Razón Social es obligatoria')
                return None, errores
            
            # Validar estado
            if estado not in ['activo', 'inactivo', 'suspendido']:
                errores.append(f'Fila {numero_fila}: Estado inválido "{estado}". Debe ser: activo, inactivo o suspendido')
                estado = 'activo'
            
            # Procesar fechas de vencimiento
            cert_venc = self._parsear_fecha(fila[3]) if len(fila) > 3 and fila[3] else None
            reso_venc = self._parsear_fecha(fila[6]) if len(fila) > 6 and fila[6] else None
            doc_venc = self._parsear_fecha(fila[9]) if len(fila) > 9 and fila[9] else None
            
            # Procesar booleanos (renovado/facturado)
            cert_renovado = self._parsear_booleano(fila[4]) if len(fila) > 4 else 0
            cert_facturado = self._parsear_booleano(fila[5]) if len(fila) > 5 else 0
            reso_renovado = self._parsear_booleano(fila[7]) if len(fila) > 7 else 0
            reso_facturado = self._parsear_booleano(fila[8]) if len(fila) > 8 else 0
            doc_renovado = self._parsear_booleano(fila[10]) if len(fila) > 10 else 0
            doc_facturado = self._parsear_booleano(fila[11]) if len(fila) > 11 else 0
            
            # Crear módulos
            from app.models.empresa import ModuloEmpresa
            from datetime import datetime
            
            certificado = ModuloEmpresa(
                activo=1 if cert_venc else 0,
                fecha_final=datetime.strptime(cert_venc, '%Y-%m-%d') if cert_venc else None,
                renovado=cert_renovado,
                facturado=cert_facturado
            )
            
            resolucion = ModuloEmpresa(
                activo=1 if reso_venc else 0,
                fecha_final=datetime.strptime(reso_venc, '%Y-%m-%d') if reso_venc else None,
                renovado=reso_renovado,
                facturado=reso_facturado
            )
            
            documento = ModuloEmpresa(
                activo=1 if doc_venc else 0,
                fecha_final=datetime.strptime(doc_venc, '%Y-%m-%d') if doc_venc else None,
                renovado=doc_renovado,
                facturado=doc_facturado
            )
            
            # Crear empresa con estructura correcta
            empresa = Empresa(
                nit=nit,
                nombre=nombre,
                tipo='Persona Jurídica',  # Valor por defecto
                estado=estado,
                certificado=certificado,
                resolucion=resolucion,
                documento=documento
            )
            
            return empresa, errores
            
        except Exception as e:
            errores.append(f'Fila {numero_fila}: Error al procesar - {str(e)}')
            return None, errores
    
    def _parsear_fecha(self, valor: Any) -> str:
        """
        Parsea un valor a formato de fecha YYYY-MM-DD
        
        Args:
            valor: Valor a parsear (puede ser datetime, string, etc.)
            
        Returns:
            Fecha en formato string o None
        """
        if not valor:
            return None
        
        try:
            # Si ya es un objeto datetime
            if isinstance(valor, datetime):
                return valor.strftime('%Y-%m-%d')
            
            # Si es string, intentar parsear
            if isinstance(valor, str):
                valor = valor.strip()
                # Intentar varios formatos
                formatos = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']
                for formato in formatos:
                    try:
                        fecha = datetime.strptime(valor, formato)
                        return fecha.strftime('%Y-%m-%d')
                    except ValueError:
                        continue
            
            return None
            
        except Exception:
            return None
    
    def _parsear_booleano(self, valor: Any) -> int:
        """
        Parsea un valor a booleano (0 o 1)
        
        Args:
            valor: Valor a parsear
            
        Returns:
            0 o 1
        """
        if not valor:
            return 0
        
        if isinstance(valor, bool):
            return 1 if valor else 0
        
        if isinstance(valor, (int, float)):
            return 1 if valor > 0 else 0
        
        if isinstance(valor, str):
            valor = valor.strip().lower()
            if valor in ['si', 'sí', 'yes', 'true', '1', 'x']:
                return 1
        
        return 0
    
    def importar_desde_excel(self, file_content: bytes) -> Dict[str, Any]:
        """
        Importa empresas desde un archivo Excel
        
        Args:
            file_content: Contenido del archivo Excel
            
        Returns:
            Diccionario con el resultado de la importación
        """
        try:
            # Validar estructura
            validacion = self.validar_estructura_excel(file_content)
            if not validacion['success']:
                return validacion
            
            # Cargar workbook
            wb = openpyxl.load_workbook(BytesIO(file_content))
            ws = wb.active
            
            # Obtener índices de columnas
            encabezados = [cell.value for cell in ws[1]]
            indices = {col: encabezados.index(col) for col in self.COLUMNAS_ESPERADAS}
            
            # Procesar filas
            resultados = {
                'total': 0,
                'exitosas': 0,
                'fallidas': 0,
                'duplicadas': 0,
                'actualizadas': 0,
                'errores': [],
                'empresas_creadas': [],
                'empresas_actualizadas': []
            }
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Saltar filas vacías
                    continue
                
                resultados['total'] += 1
                
                # Ordenar valores según columnas esperadas
                fila_ordenada = [row[indices[col]] if indices[col] < len(row) else None 
                                for col in self.COLUMNAS_ESPERADAS]
                
                # Procesar fila
                empresa, errores = self.procesar_fila(fila_ordenada, idx)
                
                if errores:
                    resultados['errores'].extend(errores)
                    resultados['fallidas'] += 1
                    continue
                
                if not empresa:
                    resultados['fallidas'] += 1
                    continue
                
                # Verificar si existe
                existe = self.repository.exists_by_nit(empresa.nit)
                
                if existe:
                    # Actualizar empresa existente
                    empresa_existente = self.repository.get_by_nit(empresa.nit)
                    empresa.id = empresa_existente.id
                    
                    success = self.repository.update(empresa)
                    if success:
                        resultados['actualizadas'] += 1
                        resultados['empresas_actualizadas'].append({
                            'nit': empresa.nit,
                            'razon_social': empresa.nombre
                        })
                    else:
                        resultados['fallidas'] += 1
                        resultados['errores'].append(
                            f'Fila {idx}: No se pudo actualizar la empresa {empresa.nit}'
                        )
                else:
                    # Crear nueva empresa
                    try:
                        empresa_creada = self.repository.create(empresa)
                        resultados['exitosas'] += 1
                        resultados['empresas_creadas'].append({
                            'nit': empresa_creada.nit,
                            'razon_social': empresa_creada.nombre
                        })
                    except Exception as e:
                        resultados['fallidas'] += 1
                        resultados['errores'].append(
                            f'Fila {idx}: Error al crear empresa - {str(e)}'
                        )
            
            return {
                'success': True,
                'message': f'Importación completada: {resultados["exitosas"]} creadas, '
                          f'{resultados["actualizadas"]} actualizadas, '
                          f'{resultados["fallidas"]} fallidas',
                'datos': resultados
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Error al importar: {str(e)}'
            }
    
    def generar_plantilla_excel(self) -> bytes:
        """
        Genera un archivo Excel de plantilla con ejemplos
        
        Returns:
            Contenido del archivo Excel en bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Empresas"
        print("Generando plantilla de Excel...")
        # Agregar encabezados
        for idx, columna in enumerate(self.COLUMNAS_ESPERADAS, start=1):
            ws.cell(row=1, column=idx, value=columna)
        
        # Agregar filas de ejemplo
        ejemplos = [
            ['900123456', 'Empresa Ejemplo S.A.S', 'activo', '2025-12-31', 'NO', 'NO', 
             '2025-06-30', 'NO', 'NO', '2025-09-15', 'NO', 'NO'],
            ['800987654', 'Comercializadora XYZ Ltda', 'activo', '2025-11-20', 'SI', 'SI', 
             '2025-08-10', 'NO', 'SI', '2025-10-05', 'SI', 'NO'],
            ['700456789', 'Industrias ABC S.A.', 'inactivo', '', '', '', 
             '', '', '', '', '', '']
        ]
        
        for idx, ejemplo in enumerate(ejemplos, start=2):
            for col_idx, valor in enumerate(ejemplo, start=1):
                ws.cell(row=idx, column=col_idx, value=valor)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
